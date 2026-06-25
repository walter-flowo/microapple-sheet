"""test_phase3.py — Phase 3 tests for liveformula and auto-detect routing.

Covers:
  1. audit_hardcoded: all-formula column → clean.
  2. audit_hardcoded: literal in formula column → hardcoded_found.
  3. audit_hardcoded: pure-input column (no formulas) → NOT flagged.
  4. audit_hardcoded: strict range_ mode → literal in range flagged.
  5. validate_formula: formula with cell ref → (True, None).
  6. validate_formula: SUM range → (True, None).
  7. validate_formula: bare number '=42' → (False, warning).
  8. validate_formula: bare expression '=2*3' → (False, warning).
  9. Auto-detect routing: bridge.is_open → True → live bridge called.
 10. Auto-detect routing: bridge.is_open → False → file engine used.
"""
from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

# Ensure src/ is on sys.path
_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
_SRC = _ROOT / "src"
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))


# ---------------------------------------------------------------------------
# audit_hardcoded tests
# ---------------------------------------------------------------------------

def test_audit_clean_all_formula_column(tmp_path: Path) -> None:
    """All-formula column → verdict clean."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A2"] = "=B2+1"
    ws["A3"] = "=B3+1"
    p = str(tmp_path / "test.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    assert result["verdict"] == "clean", f"Expected clean, got: {result}"
    assert result["hardcoded"] == [], f"Expected no hardcoded, got: {result['hardcoded']}"


def test_audit_flags_literal_in_formula_column(tmp_path: Path) -> None:
    """Formula column with one planted literal → hardcoded_found."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A2"] = "=B2+1"   # formula cell → column is a formula column
    ws["A3"] = 99.0       # numeric literal → MUST be flagged
    p = str(tmp_path / "test.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    assert result["verdict"] == "hardcoded_found", (
        f"Expected hardcoded_found, got: {result}"
    )
    cells = [h["cell"] for h in result["hardcoded"]]
    assert "A3" in cells, f"Expected A3 in hardcoded list, got: {cells}"


def test_audit_pure_input_column_not_flagged(tmp_path: Path) -> None:
    """Pure-input column (all literals, no formulas) → NOT flagged."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Input"
    ws["A2"] = 10.0
    ws["A3"] = 20.0
    ws["A4"] = 30.0
    p = str(tmp_path / "test.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    assert result["verdict"] == "clean", (
        f"Pure-input column should be clean, got: {result}"
    )
    assert result["hardcoded"] == [], f"Expected no hardcoded, got: {result['hardcoded']}"
    # The literals should be counted as input_literal_cells
    assert result["input_literal_cells"] >= 3, (
        f"Expected >=3 input_literal_cells, got: {result['input_literal_cells']}"
    )


def test_audit_strict_range_flags_literal(tmp_path: Path) -> None:
    """Strict mode: range_ given, literal in range → flagged."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = 42.0    # literal
    ws["B3"] = "=C3+1"  # formula
    p = str(tmp_path / "test.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1", range_="B2:B3")
    assert result["verdict"] == "hardcoded_found", (
        f"Expected hardcoded_found in strict mode, got: {result}"
    )
    cells = [h["cell"] for h in result["hardcoded"]]
    assert "B2" in cells, f"Expected B2 in hardcoded list, got: {cells}"
    # B3 is a formula → should NOT be in hardcoded list
    assert "B3" not in cells, f"B3 is a formula and should not be flagged, got: {cells}"


def test_audit_strict_range_clean_when_all_formulas(tmp_path: Path) -> None:
    """Strict mode: range with only formulas → clean."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["H3"] = "=I3*2"
    ws["H4"] = "=I4*2"
    ws["I3"] = "=J3+1"
    p = str(tmp_path / "test.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1", range_="H3:I4")
    assert result["verdict"] == "clean", (
        f"All-formula range should be clean, got: {result}"
    )
    assert result["hardcoded"] == []


def test_audit_mixed_sheet_flags_only_formula_column_literals(tmp_path: Path) -> None:
    """Mixed sheet: pure-input col A + formula col B with one literal → only B3 flagged."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Row 1 = headers
    ws["A1"] = "Input"
    ws["B1"] = "Output"
    # Col A = pure inputs (literals only)
    ws["A2"] = 100.0
    ws["A3"] = 200.0
    # Col B = formula column (with one planted literal)
    ws["B2"] = "=A2*2"     # formula
    ws["B3"] = 999.0       # hardcoded literal — should be flagged
    p = str(tmp_path / "test.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    assert result["verdict"] == "hardcoded_found"
    flagged_cells = [h["cell"] for h in result["hardcoded"]]
    # B3 flagged (formula column literal)
    assert "B3" in flagged_cells, f"Expected B3 flagged, got: {flagged_cells}"
    # A2, A3 NOT flagged (pure-input column)
    assert "A2" not in flagged_cells, f"A2 should NOT be flagged, got: {flagged_cells}"
    assert "A3" not in flagged_cells, f"A3 should NOT be flagged, got: {flagged_cells}"
    # Pure-input count should include A2 and A3
    assert result["input_literal_cells"] >= 2


# ---------------------------------------------------------------------------
# validate_formula tests
# ---------------------------------------------------------------------------

def test_validate_formula_with_ref() -> None:
    """Formula with cell references → (True, None)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("=B2+B3")
    assert ok is True, f"Expected True, got {ok}"
    assert warning is None, f"Expected no warning, got {warning!r}"


def test_validate_formula_sum_range() -> None:
    """SUM formula with range → (True, None)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("=SUM(A1:A10)")
    assert ok is True, f"Expected True, got {ok}"
    assert warning is None, f"Expected no warning, got {warning!r}"


def test_validate_formula_complex_nested() -> None:
    """Nested formula with multiple refs → (True, None)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("=IF(B2>0, B2*C3, 0)")
    assert ok is True, f"Expected True, got {ok}"
    assert warning is None, f"Expected no warning, got {warning!r}"


def test_validate_formula_pure_number() -> None:
    """Bare number formula '=42' → (False, warning)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("=42")
    assert ok is False, f"Expected False, got {ok}"
    assert warning is not None, "Expected a warning string"
    assert isinstance(warning, str), f"Warning should be str, got {type(warning)}"


def test_validate_formula_pure_expression() -> None:
    """Bare arithmetic '=2*3' → (False, warning)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("=2*3")
    assert ok is False, f"Expected False, got {ok}"
    assert warning is not None, "Expected a warning string"


def test_validate_formula_pure_string_literal() -> None:
    """Formula with only a string literal '=\"hello\"' → (False, warning)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula('="hello"')
    assert ok is False, f"Expected False, got {ok}"
    assert warning is not None, "Expected a warning string"


def test_validate_formula_without_leading_equals() -> None:
    """Formula without leading '=' is still processed correctly."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("B2+B3")
    assert ok is True, f"Expected True for 'B2+B3', got {ok}"


def test_validate_formula_xlookup() -> None:
    """XLOOKUP formula with cell refs → (True, None)."""
    from microapple_sheet.liveformula import validate_formula
    ok, warning = validate_formula("=XLOOKUP(A1, B1:B10, C1:C10)")
    assert ok is True, f"Expected True, got {ok}"
    assert warning is None


# ---------------------------------------------------------------------------
# Auto-detect routing tests
# ---------------------------------------------------------------------------

def test_routing_live_when_open(tmp_path: Path) -> None:
    """When bridge.is_open returns True → excel_write routes to live bridge."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "test"
    p = str(tmp_path / "open.xlsx")
    wb.save(p)

    live_result = {
        "path": p,
        "sheet": "Sheet1",
        "range": "A1",
        "cells_written": 1,
        "verified": True,
        "detail": "Unsaved in Excel",
    }

    with patch("microapple_sheet.bridge.is_open", return_value={"is_open": True}) as mock_is_open, \
         patch("microapple_sheet.bridge.live_set", return_value=live_result) as mock_live_set:
        from microapple_sheet.server import excel_write
        result = excel_write(path=p, sheet="Sheet1", range="A1", values=[[99]])

        mock_is_open.assert_called_once()
        mock_live_set.assert_called_once()
        assert result.get("routed") == "live", (
            f"Expected routed='live', got: {result.get('routed')}"
        )


def test_routing_file_when_closed(tmp_path: Path) -> None:
    """When bridge.is_open returns False → excel_write routes to file engine."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "test"
    p = str(tmp_path / "closed.xlsx")
    wb.save(p)

    with patch("microapple_sheet.bridge.is_open", return_value={"is_open": False}):
        from microapple_sheet.server import excel_write
        result = excel_write(path=p, sheet="Sheet1", range="A1", values=[[42]])
        assert result.get("routed") == "file", (
            f"Expected routed='file', got: {result.get('routed')}"
        )
        # Verify the value was actually written
        from microapple_sheet.ops_openpyxl import read_range
        r = read_range(p, "Sheet1", "A1", mode="values")
        assert r["data"][0][0] == 42


def test_routing_set_cell_live(tmp_path: Path) -> None:
    """excel_set_cell routes to live bridge when file is open."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    p = str(tmp_path / "open2.xlsx")
    wb.save(p)

    live_result = {
        "path": p, "sheet": "Sheet1", "range": "B4",
        "cells_written": 1, "verified": True, "detail": "Unsaved",
    }

    with patch("microapple_sheet.bridge.is_open", return_value={"is_open": True}), \
         patch("microapple_sheet.bridge.live_set", return_value=live_result):
        from microapple_sheet.server import excel_set_cell
        result = excel_set_cell(path=p, sheet="Sheet1", cell="B4", value=99)
        assert result.get("routed") == "live"


def test_routing_write_formula_live_with_warning(tmp_path: Path) -> None:
    """excel_write_formula: live routing + formula_warning when formula is a bare constant."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    p = str(tmp_path / "formula_warn.xlsx")
    wb.save(p)

    live_result = {
        "path": p, "sheet": "Sheet1", "cell": "C3",
        "formula": "=42", "verified_value": 42, "detail": "Unsaved",
    }

    with patch("microapple_sheet.bridge.is_open", return_value={"is_open": True}), \
         patch("microapple_sheet.bridge.live_formula", return_value=live_result):
        from microapple_sheet.server import excel_write_formula
        result = excel_write_formula(path=p, sheet="Sheet1", cell="C3", formula="=42")
        assert result.get("routed") == "live"
        assert "formula_warning" in result, (
            "Expected formula_warning for bare constant '=42'"
        )


def test_routing_write_formula_file_no_warning(tmp_path: Path) -> None:
    """excel_write_formula: file routing + no warning when formula references cells."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    p = str(tmp_path / "formula_ok.xlsx")
    wb.save(p)

    with patch("microapple_sheet.bridge.is_open", return_value={"is_open": False}):
        from microapple_sheet.server import excel_write_formula
        result = excel_write_formula(
            path=p, sheet="Sheet1", cell="C3", formula="=B2+B3"
        )
        assert result.get("routed") == "file"
        assert "formula_warning" not in result, (
            f"No warning expected for valid formula, got: {result.get('formula_warning')}"
        )


# ---------------------------------------------------------------------------
# compose_linked_formula tests
# ---------------------------------------------------------------------------

def test_compose_linked_formula_basic() -> None:
    """compose_linked_formula substitutes placeholders correctly."""
    from microapple_sheet.liveformula import compose_linked_formula
    result = compose_linked_formula("={peak}*{fraction}", {"peak": "B4", "fraction": "$B$2"})
    assert result == "=B4*$B$2", f"Got: {result!r}"


def test_compose_linked_formula_ceiling() -> None:
    """compose_linked_formula works with CEILING-style templates."""
    from microapple_sheet.liveformula import compose_linked_formula
    result = compose_linked_formula("=CEILING({raw},50)", {"raw": "B6"})
    assert result == "=CEILING(B6,50)", f"Got: {result!r}"


# ---------------------------------------------------------------------------
# Totals-row aware audit tests (phase 3 bug-fix)
# ---------------------------------------------------------------------------

def test_audit_input_column_with_sum_total_not_flagged(tmp_path: Path) -> None:
    """Input column with a bottom SUM total → inputs NOT flagged."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Row 1: headers
    ws["A1"] = "Item"
    ws["B1"] = "Value"
    # Data rows: B is an input column (all literals)
    ws["A2"] = "Item 1"
    ws["B2"] = 100.0
    ws["A3"] = "Item 2"
    ws["B3"] = 200.0
    # Totals row: SUM formula
    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"
    p = str(tmp_path / "test_sum.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    # B2 and B3 are inputs, must NOT be flagged
    assert result["verdict"] == "clean", f"Expected clean, got: {result}"
    cells = [h["cell"] for h in result["hardcoded"]]
    assert "B2" not in cells, f"B2 must not be flagged; got: {cells}"
    assert "B3" not in cells, f"B3 must not be flagged; got: {cells}"


def test_audit_hardcoded_total_in_totals_row_flagged(tmp_path: Path) -> None:
    """Numeric literal typed over a totals row → flagged as hardcoded total."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Item"
    ws["B1"] = "Calc"
    # Data rows: B has a formula (formula column)
    ws["A2"] = "Item 1"
    ws["B2"] = "=C2*2"
    ws["A3"] = "Item 2"
    ws["B3"] = "=C3*2"
    # Totals row: B has a hardcoded number instead of a formula
    ws["A4"] = "Total"
    ws["B4"] = 999.0   # should be =SUM(B2:B3) — this is the bug
    p = str(tmp_path / "test_hardcoded_total.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    assert result["verdict"] == "hardcoded_found", f"Expected hardcoded_found, got: {result}"
    cells = [h["cell"] for h in result["hardcoded"]]
    assert "B4" in cells, f"Expected B4 in hardcoded list, got: {cells}"
    # Check reason mentions "total"
    b4_entry = next(h for h in result["hardcoded"] if h["cell"] == "B4")
    assert "total" in b4_entry["reason"].lower(), (
        f"Expected 'total' in reason, got: {b4_entry['reason']!r}"
    )


def test_audit_data_row_literal_in_formula_column_still_flagged(tmp_path: Path) -> None:
    """Data-row literal in a formula column still flagged (core heuristic still works)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["B1"] = "Result"
    # Data rows: B has a formula in row 2, literal in row 3
    ws["A2"] = "R1"
    ws["B2"] = "=C2*1.1"   # formula → B is a formula column
    ws["A3"] = "R2"
    ws["B3"] = 50.0         # literal in a formula column → MUST flag
    # Totals row: proper SUM formula (should NOT flag)
    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"
    p = str(tmp_path / "test_data_literal.xlsx")
    wb.save(p)

    from microapple_sheet.liveformula import audit_hardcoded
    result = audit_hardcoded(p, "Sheet1")
    assert result["verdict"] == "hardcoded_found", f"Expected hardcoded_found, got: {result}"
    cells = [h["cell"] for h in result["hardcoded"]]
    assert "B3" in cells, f"Expected B3 flagged; got: {cells}"   # data-row literal flagged
    assert "B4" not in cells, f"B4 is a SUM formula and must not be flagged; got: {cells}"
