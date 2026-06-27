"""ops_openpyxl.py — read/write file-engine operations via openpyxl.

Phase 1b — IMPLEMENTED.

READ operations are safe for any workbook state (read-only, no modification risk).
WRITE operations are engine-routed via classify_workbook() to preserve Rule-0 safety:

  Engine.OPENPYXL  (new / safe — no formula-cached <v> values)
      Direct openpyxl load → modify → save. No caches to lose.

  Engine.LIBREOFFICE  (computed — formula-cached <v> values present)
      Value/formula writes: libreoffice.apply_edits() [openpyxl write → LO recalc].
      Structural changes (format, define-name): openpyxl modify → libreoffice.recalc().
      CAVEAT: x14 CF (databars/icon-sets/colour-scales) may be corrupted by the
      openpyxl round-trip. Plain dxf cellIs/expression CF survives. See apply_edits
      docstring. Post-MVP: use LO-macro setValue path for value-only edits.

  Engine.XML_SURGERY  (dynamic-array / LAMBDA)
      Refuse all file-mode writes. Use excel_live_set / excel_live_formula (Phase 2).

  Engine.EXCEL_LIVE  (workbook open in Excel)
      Refuse all file-mode writes. Use live-bridge tools.

Clobber-safety: every write checks is_open_in_excel() (lock-file heuristic, Phase 1).
Backups: every write calls backups.snapshot() before touching the file.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.workbook.defined_name import DefinedName

from microapple_sheet import backups
from microapple_sheet.engine import Engine, WorkbookClass, classify_workbook, is_open_in_excel


# ---------------------------------------------------------------------------
# Guards & helpers
# ---------------------------------------------------------------------------

def _clobber_guard(path: Path) -> None:
    """Raise RuntimeError if Excel has the workbook locked."""
    if is_open_in_excel(path):
        raise RuntimeError(
            f"Excel lock file (~${path.name}) detected — workbook is open in Excel. "
            "Close it first or use excel_live_set / excel_live_formula (Phase 2)."
        )


def _write_guard(cls: WorkbookClass, path: Path) -> None:
    """Raise for engines that cannot accept file-mode writes."""
    if cls.engine == Engine.XML_SURGERY:
        raise ValueError(
            f"{path.name} contains dynamic-array / LAMBDA formulas. "
            "File-mode writes are unsafe (LO recalc cascades _xlfn formulas to #N/A). "
            "Open in Excel and use excel_live_set / excel_live_formula instead."
        )
    if cls.engine == Engine.EXCEL_LIVE:
        raise RuntimeError(
            f"{path.name} is open in Excel — use live-bridge tools "
            "(excel_live_set / excel_live_formula) for live edits."
        )


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Return (min_row, min_col, max_row, max_col) 1-based from A1 range string."""
    s = range_str.upper()
    if ":" in s:
        a, b = s.split(":", 1)
        sr, sc = coordinate_to_tuple(a)
        er, ec = coordinate_to_tuple(b)
        return min(sr, er), min(sc, ec), max(sr, er), max(sc, ec)
    r, c = coordinate_to_tuple(s)
    return r, c, r, c


def _values_to_edits(
    sheet: str, min_row: int, min_col: int, values: list[list[Any]]
) -> list[dict[str, Any]]:
    """Flatten a 2D values array into apply_edits edit-dict list."""
    edits: list[dict[str, Any]] = []
    for r_idx, row in enumerate(values):
        for c_idx, val in enumerate(row):
            cell_ref = f"{get_column_letter(min_col + c_idx)}{min_row + r_idx}"
            if isinstance(val, str) and val.startswith("="):
                edits.append({"sheet": sheet, "cell": cell_ref, "type": "formula", "data": val})
            else:
                edits.append({"sheet": sheet, "cell": cell_ref, "type": "value", "data": val})
    return edits


def _openpyxl_then_recalc(
    path: Path,
    cls: WorkbookClass,
    mutate,
    backup_path: str,
) -> dict[str, Any]:
    """Apply mutate(wb) via openpyxl then, if needed, run LO recalc.

    Used for structural changes (format, define-name) on computed workbooks.
    mutate receives the openpyxl Workbook and returns an optional extra-fields dict.
    """
    wb = openpyxl.load_workbook(str(path))
    extra: dict[str, Any] = {}
    try:
        extra = mutate(wb) or {}
        wb.save(str(path))
    finally:
        wb.close()

    recalc_info: dict[str, Any] = {}
    if cls.engine == Engine.LIBREOFFICE:
        from microapple_sheet import libreoffice as lo
        r = lo.recalc(path)
        recalc_info = {"recalc_ok": r["recalc_ok"], "recalc_detail": r["detail"]}
    else:
        recalc_info = {"recalc_ok": True, "recalc_detail": "no recalc needed"}

    return {
        "path": str(path),
        "backup_path": backup_path,
        "engine": cls.engine.value,
        **extra,
        **recalc_info,
    }


# ---------------------------------------------------------------------------
# Create
# ---------------------------------------------------------------------------

def create(path: str | Path, sheets: list[str] | None = None) -> dict[str, Any]:
    """Create a fresh .xlsx workbook at *path*.

    Args:
        path:   Destination path (parent directory must exist).
        sheets: Sheet names in order. Default: ['Sheet1'].

    Returns:
        {path, sheets, engine: 'openpyxl'}
    """
    p = Path(path)
    names = sheets or ["Sheet1"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = names[0]
    for name in names[1:]:
        wb.create_sheet(name)

    wb.save(str(p))
    wb.close()

    return {"path": str(p), "sheets": names, "engine": "openpyxl"}


# ---------------------------------------------------------------------------
# Read (no engine routing — read-only, safe for any workbook)
# ---------------------------------------------------------------------------

def read_range(
    path: str | Path,
    sheet: str,
    range_: str | None = None,
    mode: str = "values",
) -> dict[str, Any]:
    """Read cells from a worksheet.

    Args:
        path:   Absolute path to the workbook.
        sheet:  Sheet name.
        range_: A1-notation range (e.g. 'B2:D10'). None → entire used range.
        mode:   'values' | 'formulas' | 'both'.
            'values'  — data_only=True; reads cached <v> (None if not yet recalced).
            'formulas'— data_only=False; reads formula strings (e.g. '=SUM(A1:A10)').
            'both'    — returns both keys: 'values' and 'formulas'.

    Returns:
        {sheet, range, mode, data: list[list[Any]]}
        For mode='both': {sheet, range, mode, values: 2D, formulas: 2D}
    """
    p = Path(path)
    mode = mode.lower()

    if mode == "both":
        val_result = read_range(p, sheet, range_, mode="values")
        frm_result = read_range(p, sheet, range_, mode="formulas")
        return {
            "sheet": sheet,
            "range": val_result["range"],
            "mode": "both",
            "values": val_result["data"],
            "formulas": frm_result["data"],
        }

    data_only = mode != "formulas"
    wb = openpyxl.load_workbook(str(p), data_only=data_only, read_only=True)
    try:
        ws = wb[sheet]

        if range_:
            min_r, min_c, max_r, max_c = _parse_range(range_)
            range_used = range_.upper()
        else:
            min_r = ws.min_row or 1
            max_r = ws.max_row or 1
            min_c = ws.min_column or 1
            max_c = ws.max_column or 1
            range_used = (
                f"{get_column_letter(min_c)}{min_r}:"
                f"{get_column_letter(max_c)}{max_r}"
            )

        data: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_r, max_row=max_r,
            min_col=min_c, max_col=max_c,
            values_only=True,
        ):
            data.append(list(row))

    finally:
        wb.close()

    return {"sheet": sheet, "range": range_used, "mode": mode, "data": data}


def read_table(
    path: str | Path,
    sheet: str,
    range_: str | None = None,
) -> dict[str, Any]:
    """Read a worksheet table as a list of header-keyed dicts.

    The first row is treated as column headers. Empty header cells are named
    'Col_<column_letter>'. Trailing None cells in data rows are preserved.

    Args:
        path:   Absolute path to the workbook.
        sheet:  Sheet name.
        range_: A1 range to read. None → entire used range.

    Returns:
        {sheet, range, headers: list[str], records: list[dict[str, Any]], rows: int}
    """
    raw = read_range(path, sheet, range_=range_, mode="values")
    data = raw["data"]
    if not data:
        return {"sheet": sheet, "range": raw["range"], "headers": [], "records": [], "rows": 0}

    header_row = data[0]
    headers = [
        (str(h) if h is not None else f"Col_{get_column_letter(i + 1)}")
        for i, h in enumerate(header_row)
    ]
    records = [dict(zip(headers, row)) for row in data[1:]]
    return {
        "sheet": sheet,
        "range": raw["range"],
        "headers": headers,
        "records": records,
        "rows": len(records),
    }


def list_sheets(path: str | Path) -> dict[str, Any]:
    """Return ordered sheet names for a workbook.

    Args:
        path: Absolute path to the workbook.

    Returns:
        {path, sheets: list[str]}
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True)
    try:
        sheets = wb.sheetnames
    finally:
        wb.close()
    return {"path": str(p), "sheets": sheets}


def list_names(path: str | Path) -> dict[str, Any]:
    """Return all defined names in the workbook.

    Args:
        path: Absolute path to the workbook.

    Returns:
        {path, names: list[{name, ref, scope}]}
    """
    p = Path(path)
    # Use data_only=False (read_only=True may strip defined_names in some LO-saved files)
    wb = openpyxl.load_workbook(str(p), read_only=False, data_only=True)
    try:
        names: list[dict[str, Any]] = []
        for name_key in wb.defined_names:
            dn = wb.defined_names[name_key]
            scope = dn.localSheetId
            sheet_scope = (
                "workbook"
                if scope is None
                else (wb.worksheets[scope].title if scope < len(wb.worksheets) else f"sheet_{scope}")
            )
            names.append({
                "name": dn.name,
                "ref": dn.attr_text or "",
                "scope": sheet_scope,
            })
    finally:
        wb.close()
    return {"path": str(p), "names": names}


# ---------------------------------------------------------------------------
# Writes (engine-routed)
# ---------------------------------------------------------------------------

def write_range(
    path: str | Path,
    sheet: str,
    range_: str,
    values: list[list[Any]],
) -> dict[str, Any]:
    """Write a 2-D array of values (or formulas) to a range, engine-routed.

    Values that are strings starting with '=' are treated as formulas.

    Args:
        path:   Absolute path to the workbook.
        sheet:  Sheet name.
        range_: Top-left cell or full A1 range (e.g. 'B4' or 'B4:D10').
        values: Row-major 2-D list. Use None for empty cells.

    Returns:
        {path, sheet, range, engine, cells_written, backup_path, recalc_ok}
    """
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)

    min_r, min_c, _, _ = _parse_range(range_)
    cells_written = sum(len(row) for row in values)

    if cls.engine == Engine.LIBREOFFICE:
        edits = _values_to_edits(sheet, min_r, min_c, values)
        from microapple_sheet import libreoffice as lo
        result = lo.apply_edits(p, edits)
        result["sheet"] = sheet
        result["range"] = range_
        result["cells_written"] = cells_written
        return result

    # OPENPYXL engine — direct write
    backup_path = backups.snapshot(p)
    wb = openpyxl.load_workbook(str(p))
    try:
        ws = wb[sheet]
        for r_idx, row in enumerate(values):
            for c_idx, val in enumerate(row):
                ws.cell(row=min_r + r_idx, column=min_c + c_idx, value=val)
        wb.save(str(p))
    finally:
        wb.close()

    return {
        "path": str(p),
        "sheet": sheet,
        "range": range_,
        "engine": "openpyxl",
        "cells_written": cells_written,
        "backup_path": backup_path,
        "recalc_ok": True,
    }


def set_cell(
    path: str | Path,
    sheet: str,
    cell: str,
    value: Any,
) -> dict[str, Any]:
    """Write a single cell value (engine-routed convenience wrapper).

    Args:
        path:  Absolute path to the workbook.
        sheet: Sheet name.
        cell:  A1 cell reference (e.g. 'B4').
        value: Literal value or formula string.

    Returns:
        {path, sheet, cell, engine, backup_path, recalc_ok}
    """
    result = write_range(path, sheet, cell, [[value]])
    result["cell"] = cell
    result.pop("range", None)
    result.pop("cells_written", None)
    return result


def write_formula(
    path: str | Path,
    sheet: str,
    cell: str,
    formula: str,
) -> dict[str, Any]:
    """Write a single formula to a cell (engine-routed).

    Formula validation (liveformula module) is Phase 3. For now the formula
    is written as-is after normalising the leading '='.

    Args:
        path:    Absolute path to the workbook.
        sheet:   Sheet name.
        cell:    A1 cell reference (e.g. 'B4').
        formula: Formula string — with or without leading '='.

    Returns:
        {path, sheet, cell, formula, engine, backup_path, recalc_ok}
    """
    norm = formula.strip()
    if not norm.startswith("="):
        norm = "=" + norm

    result = write_range(path, sheet, cell, [[norm]])
    result["cell"] = cell
    result["formula"] = norm
    result.pop("range", None)
    result.pop("cells_written", None)
    return result


def write_linked_formula(
    path: str | Path,
    sheet: str,
    cell: str,
    template: str,
    refs: dict[str, str],
) -> dict[str, Any]:
    """Compose and write a formula by substituting cell-coordinate references.

    Implements Rule #1/#4 (live-formula / linked-formula principle): the caller
    supplies captured cell coordinates (from Cell.coordinate) rather than
    hardcoded strings, so the formula is always live and traceable.

    Args:
        path:     Absolute path to the workbook.
        sheet:    Sheet name.
        cell:     Target cell in A1 notation.
        template: Formula template with {name} placeholders.
                  Example: '={base}*{rate}'
        refs:     Mapping of placeholder name → A1 coordinate.
                  Example: {'base': 'B4', 'rate': '$B$2'}

    Returns:
        {path, sheet, cell, resolved_formula, engine, backup_path, recalc_ok}

    Example:
        write_linked_formula(path, 'Calcs', 'D4',
                             '={peak}*{fraction}',
                             {'peak': 'B4', 'fraction': '$B$2'})
        → writes formula '=B4*$B$2' to Calcs!D4
    """
    resolved = template.format(**refs)
    result = write_formula(path, sheet, cell, resolved)
    result["resolved_formula"] = result.pop("formula", resolved)
    return result


def format_range(
    path: str | Path,
    sheet: str,
    range_: str,
    *,
    font: dict[str, Any] | None = None,
    fill: dict[str, Any] | None = None,
    border: dict[str, Any] | None = None,
    number_format: str | None = None,
    alignment: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Apply formatting to a cell range (engine-routed).

    For LIBREOFFICE-engine workbooks, applies formatting via openpyxl then
    immediately runs LO recalc to restore cached formula values. Note: x14
    conditional formatting (databars, icon-sets) may be corrupted by the
    openpyxl round-trip; plain dxf cellIs / expression CF survives.

    Args:
        path:          Absolute path to the workbook.
        sheet:         Sheet name.
        range_:        A1-notation range.
        font:          {name?, size?, bold?, italic?, color?} (color = hex RRGGBB).
        fill:          {fgColor?} (hex RRGGBB; solid PatternFill).
        border:        {top?, bottom?, left?, right?} each {style?, color?}.
        number_format: Excel number-format string (e.g. '#,##0.00').
        alignment:     {horizontal?, vertical?, wrap_text?}.

    Returns:
        {path, sheet, range, engine, backup_path, recalc_ok, recalc_detail}
    """
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)

    backup_path = backups.snapshot(p)
    min_r, min_c, max_r, max_c = _parse_range(range_)

    def mutate(wb: openpyxl.Workbook) -> dict[str, Any]:
        ws = wb[sheet]

        _font = Font(**{k: v for k, v in (font or {}).items()}) if font else None
        _fill = (
            PatternFill(
                fill_type="solid",
                fgColor=fill.get("fgColor", "FFFFFF"),
            )
            if fill
            else None
        )
        _alignment = Alignment(**{k: v for k, v in (alignment or {}).items()}) if alignment else None

        _border: Border | None = None
        if border:
            sides = {}
            for side_name in ("top", "bottom", "left", "right"):
                sd = border.get(side_name)
                if sd:
                    sides[side_name] = Side(
                        border_style=sd.get("style", "thin"),
                        color=sd.get("color", "000000"),
                    )
            _border = Border(**sides) if sides else None

        for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
            for cell in row:
                if _font:
                    cell.font = _font
                if _fill:
                    cell.fill = _fill
                if _border:
                    cell.border = _border
                if _alignment:
                    cell.alignment = _alignment
                if number_format:
                    cell.number_format = number_format

        return {"sheet": sheet, "range": range_}

    return _openpyxl_then_recalc(p, cls, mutate, backup_path)


def define_name(
    path: str | Path,
    name: str,
    ref: str,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Add or update a defined name / named range in the workbook.

    For LIBREOFFICE-engine workbooks, runs LO recalc after saving to restore
    cached formula values.

    Args:
        path:  Absolute path to the workbook.
        name:  Defined name (must be a valid Excel identifier, no spaces).
        ref:   Reference in A1 notation (e.g. 'Sheet1!$B$4:$B$10' or '42').
        sheet: If given, scope the name to this sheet (localSheetId).

    Returns:
        {path, name, ref, scope, engine, backup_path, recalc_ok}
    """
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)

    backup_path = backups.snapshot(p)

    def mutate(wb: openpyxl.Workbook) -> dict[str, Any]:
        # Determine localSheetId if sheet-scoped
        local_sheet_id: int | None = None
        if sheet:
            try:
                local_sheet_id = wb.sheetnames.index(sheet)
            except ValueError:
                raise ValueError(f"Sheet {sheet!r} not found in workbook.") from None

        # Remove existing definition if present (handles both workbook + sheet scopes)
        if name in wb.defined_names:
            wb.defined_names.pop(name)

        dn = DefinedName(name=name, attr_text=ref)
        if local_sheet_id is not None:
            dn.localSheetId = local_sheet_id
        wb.defined_names.add(dn)

        scope = sheet if sheet else "workbook"
        return {"name": name, "ref": ref, "scope": scope}

    return _openpyxl_then_recalc(p, cls, mutate, backup_path)


# ---------------------------------------------------------------------------
# Sheet management (file engine) — add / delete / move / rename
# ---------------------------------------------------------------------------

def add_sheet(
    path: str | Path, name: str, index: int | None = None
) -> dict[str, Any]:
    """Add a worksheet to a closed workbook (engine-routed, Rule-0 safe)."""
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)
    backup_path = backups.snapshot(p)

    def mutate(wb: openpyxl.Workbook) -> dict[str, Any]:
        if name in wb.sheetnames:
            raise ValueError(f"Sheet {name!r} already exists.")
        wb.create_sheet(title=name, index=index)
        return {"sheet": name, "sheets": wb.sheetnames}

    return _openpyxl_then_recalc(p, cls, mutate, backup_path)


def delete_sheet(path: str | Path, name: str) -> dict[str, Any]:
    """Delete a worksheet from a closed workbook (engine-routed, Rule-0 safe)."""
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)
    backup_path = backups.snapshot(p)

    def mutate(wb: openpyxl.Workbook) -> dict[str, Any]:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet {name!r} not found.")
        if len(wb.sheetnames) <= 1:
            raise ValueError("Cannot delete the only sheet in a workbook.")
        del wb[name]
        return {"sheet": name, "sheets": wb.sheetnames}

    return _openpyxl_then_recalc(p, cls, mutate, backup_path)


def rename_sheet(
    path: str | Path, old_name: str, new_name: str
) -> dict[str, Any]:
    """Rename a worksheet in a closed workbook (engine-routed, Rule-0 safe)."""
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)
    backup_path = backups.snapshot(p)

    def mutate(wb: openpyxl.Workbook) -> dict[str, Any]:
        if old_name not in wb.sheetnames:
            raise ValueError(f"Sheet {old_name!r} not found.")
        if new_name in wb.sheetnames:
            raise ValueError(f"Sheet {new_name!r} already exists.")
        wb[old_name].title = new_name
        return {"old_name": old_name, "new_name": new_name, "sheets": wb.sheetnames}

    return _openpyxl_then_recalc(p, cls, mutate, backup_path)


def move_sheet(
    path: str | Path,
    name: str,
    before: str | None = None,
    after: str | None = None,
) -> dict[str, Any]:
    """Move a worksheet before/after an anchor sheet in a closed workbook."""
    p = Path(path)
    _clobber_guard(p)
    cls = classify_workbook(p)
    _write_guard(cls, p)
    backup_path = backups.snapshot(p)

    def mutate(wb: openpyxl.Workbook) -> dict[str, Any]:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet {name!r} not found.")
        anchor = after if after else before
        if not anchor:
            raise ValueError("move_sheet requires before= or after=.")
        if anchor not in wb.sheetnames:
            raise ValueError(f"Anchor sheet {anchor!r} not found.")
        ws = wb[name]
        wb._sheets.remove(ws)  # noqa: SLF001 — stable openpyxl reorder idiom
        insert_at = wb.sheetnames.index(anchor) + (1 if after else 0)
        wb._sheets.insert(insert_at, ws)  # noqa: SLF001
        return {"sheet": name, "sheets": wb.sheetnames}

    return _openpyxl_then_recalc(p, cls, mutate, backup_path)
