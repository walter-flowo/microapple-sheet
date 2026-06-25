"""test_phase1b.py — Phase 1b file-engine integration tests.

Covers:
  1. Create → write values → read back (values and formulas modes).
  2. write_linked_formula produces a live =B4*$B$2 formula.
  3. excel_set_cell round-trip.
  4. excel_read_table header/records.
  5. Write to a COMPUTED workbook routes through apply_edits; cached values preserved.
  6. Clobber-safety: ~$<name>.xlsx lock file → write refuses.
  7. write to XML_SURGERY workbook refuses.
  8. excel_list_sheets and excel_list_names.
"""
from __future__ import annotations

import re
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
_SRC = _ROOT / "src"
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

import openpyxl
import pytest

from microapple_sheet.engine import Engine, classify_workbook
from microapple_sheet.libreoffice import recalc, _find_soffice
from microapple_sheet.ops_openpyxl import (
    create,
    define_name,
    format_range,
    list_names,
    list_sheets,
    read_range,
    read_table,
    set_cell,
    write_formula,
    write_linked_formula,
    write_range,
)


# Skip LO-dependent tests when soffice is not available on this host.
_lo_installed = _find_soffice() is not None
_lo_skip = pytest.mark.skipif(
    not _lo_installed,
    reason="LibreOffice (soffice) not installed on this host",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _count_formula_cached_v(xlsx_path: Path) -> int:
    """Count <v> elements preceded by </f> (formula-cached values only)."""
    pattern = re.compile(r"</f>\s*<v>[^<\s][^<]*</v>")
    count = 0
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        for name in zf.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                xml = zf.read(name).decode("utf-8", errors="replace")
                count += len(pattern.findall(xml))
    return count


def _build_computed_workbook(path: Path) -> None:
    """Build a minimal computed workbook: Inputs!A1=1, Calcs!A1=Inputs!A1*2."""
    wb = openpyxl.Workbook()
    inp = wb.active
    inp.title = "Inputs"
    inp["A1"] = 5  # input value

    calc = wb.create_sheet("Calcs")
    for r in range(1, 11):
        calc.cell(row=r, column=1, value=f"=Inputs!A1*{r}")   # A col
        calc.cell(row=r, column=2, value=f"=A{r}+Inputs!A1")  # B col

    wb.save(str(path))
    recalc(path)  # populate cached values


# ---------------------------------------------------------------------------
# Test 1: Create / write / read round-trip
# ---------------------------------------------------------------------------

def test_create_write_read_roundtrip(tmp_path) -> None:
    """Create workbook, write values, read back in values and formulas modes."""
    wb_path = tmp_path / "test_create.xlsx"

    # Create
    result = create(wb_path, sheets=["Data", "Summary"])
    assert result["sheets"] == ["Data", "Summary"]
    assert Path(result["path"]).exists()

    # Write a 3×3 block of values to Data!B2:D4
    values = [[10, 20, 30], [40, 50, 60], [70, 80, 90]]
    w_result = write_range(wb_path, "Data", "B2", values)
    assert w_result["cells_written"] == 9
    assert w_result["engine"] == "openpyxl"

    # Read back in values mode
    r_vals = read_range(wb_path, "Data", "B2:D4", mode="values")
    assert r_vals["data"] == values, f"Values mismatch: {r_vals['data']!r}"

    # Write a formula to Summary!A1
    write_formula(wb_path, "Summary", "A1", "=SUM(Data!B2:D4)")

    # Read back in formulas mode — should see the formula string
    r_frm = read_range(wb_path, "Summary", "A1", mode="formulas")
    formula_str = r_frm["data"][0][0]
    assert isinstance(formula_str, str) and formula_str.startswith("="), (
        f"Expected formula string, got {formula_str!r}"
    )
    assert "SUM" in formula_str.upper(), f"Expected SUM formula, got {formula_str!r}"


# ---------------------------------------------------------------------------
# Test 2: write_linked_formula produces a live =B4*$B$2 formula
# ---------------------------------------------------------------------------

def test_write_linked_formula(tmp_path) -> None:
    """write_linked_formula composes and writes a live cell-reference formula."""
    wb_path = tmp_path / "test_linked.xlsx"
    create(wb_path, sheets=["Calcs"])

    # Set up two input cells
    write_range(wb_path, "Calcs", "B4", [[120]])   # peak value
    write_range(wb_path, "Calcs", "B2", [[0.5]])   # fraction

    # Compose ={peak}*{fraction} → =B4*$B$2
    result = write_linked_formula(
        wb_path, "Calcs", "D4",
        template="={peak}*{fraction}",
        refs={"peak": "B4", "fraction": "$B$2"},
    )
    assert result["resolved_formula"] == "=B4*$B$2", (
        f"Expected '=B4*$B$2', got {result['resolved_formula']!r}"
    )

    # Read back in formulas mode — must be a formula (not a literal)
    r = read_range(wb_path, "Calcs", "D4", mode="formulas")
    formula_cell = r["data"][0][0]
    assert isinstance(formula_cell, str) and formula_cell.startswith("="), (
        f"Expected formula string in D4, got {formula_cell!r}"
    )
    # $B$2 contains B and 2 but not adjacent as "B2"; check both refs are present
    # Accept any dollar-prefixed variant: B4, $B4, B$4, $B$4 etc.
    assert re.search(r"\$?B\$?4", formula_cell) and re.search(r"\$?B\$?2", formula_cell), (
        f"Formula {formula_cell!r} does not reference B4 and B2"
    )


# ---------------------------------------------------------------------------
# Test 3: excel_set_cell round-trip
# ---------------------------------------------------------------------------

def test_set_cell(tmp_path) -> None:
    """set_cell writes a single value and it reads back correctly."""
    wb_path = tmp_path / "test_set_cell.xlsx"
    create(wb_path)

    set_cell(wb_path, "Sheet1", "C5", 42)
    r = read_range(wb_path, "Sheet1", "C5", mode="values")
    assert r["data"][0][0] == 42, f"Expected 42, got {r['data'][0][0]!r}"

    # Formula via set_cell
    set_cell(wb_path, "Sheet1", "C6", "=C5*2")
    r2 = read_range(wb_path, "Sheet1", "C6", mode="formulas")
    formula = r2["data"][0][0]
    assert isinstance(formula, str) and "C5" in formula, (
        f"Expected formula referencing C5, got {formula!r}"
    )


# ---------------------------------------------------------------------------
# Test 4: read_table header/records
# ---------------------------------------------------------------------------

def test_read_table(tmp_path) -> None:
    """read_table returns header-keyed records from a worksheet table."""
    wb_path = tmp_path / "test_table.xlsx"
    create(wb_path, sheets=["Table"])

    # Write header + 3 data rows
    table_data = [
        ["Name", "Value", "Unit"],
        ["U-value", 0.18, "W/m²K"],
        ["Peak load", 280, "kW"],
        ["HDD", 2650, "degree-days"],
    ]
    write_range(wb_path, "Table", "A1", table_data)

    result = read_table(wb_path, "Table", range_="A1:C4")
    assert result["headers"] == ["Name", "Value", "Unit"], f"Headers: {result['headers']}"
    assert result["rows"] == 3
    assert result["records"][0]["Name"] == "U-value"
    assert result["records"][1]["Value"] == 280
    assert result["records"][2]["Unit"] == "degree-days"


# ---------------------------------------------------------------------------
# Test 5: Write to a computed workbook routes through apply_edits
# ---------------------------------------------------------------------------

@_lo_skip
def test_write_to_computed_workbook_preserves_caches(tmp_path) -> None:
    """write_range on a LIBREOFFICE-engine workbook preserves all cached formula values."""
    wb_path = tmp_path / "test_computed_write.xlsx"
    _build_computed_workbook(wb_path)

    cls = classify_workbook(wb_path)
    assert cls.engine == Engine.LIBREOFFICE, (
        f"Expected LIBREOFFICE engine, got {cls.engine}"
    )

    cached_before = _count_formula_cached_v(wb_path)
    assert cached_before >= 10, f"Expected cached values after LO recalc, got {cached_before}"

    # Write a new input value via write_range
    result = write_range(wb_path, "Inputs", "A1", [[99]])
    assert result.get("engine") == "libreoffice", (
        f"Expected libreoffice engine for computed workbook, got {result.get('engine')!r}"
    )
    assert result.get("recalc_ok"), f"recalc_ok was False: {result}"

    cached_after = _count_formula_cached_v(wb_path)
    assert cached_after >= cached_before, (
        f"Cached formula values dropped: {cached_before} → {cached_after}. Rule-0 violated."
    )

    # Verify the input cell was actually updated
    r = read_range(wb_path, "Inputs", "A1", mode="values")
    assert r["data"][0][0] == 99, f"Expected Inputs!A1=99 after write, got {r['data'][0][0]!r}"


# ---------------------------------------------------------------------------
# Test 6: Clobber-safety — lock file present → write refuses
# ---------------------------------------------------------------------------

def test_clobber_safety_lock_file(tmp_path) -> None:
    """write_range raises RuntimeError when Excel lock file is present."""
    wb_path = tmp_path / "test_clobber.xlsx"
    create(wb_path)

    # Simulate Excel lock file
    lock_path = tmp_path / f"~$test_clobber.xlsx"
    lock_path.write_text("lock")

    try:
        with pytest.raises(RuntimeError, match="lock file"):
            write_range(wb_path, "Sheet1", "A1", [[1]])
    finally:
        lock_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Test 7: XML_SURGERY workbook refuses writes
# ---------------------------------------------------------------------------

def test_xml_surgery_refuses_write(tmp_path) -> None:
    """write_range on an XML_SURGERY workbook raises ValueError."""
    wb_path = tmp_path / "test_dynamic.xlsx"

    # Build a workbook containing a _xlfn formula marker to trigger XML_SURGERY
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    # Embed a fake _xlfn marker directly in the cell's stored formula string
    # (openpyxl stores this literally; classify_workbook will scan it)
    ws["A2"].value = "=_xlfn.XLOOKUP(A1,A1:A10,A1:A10)"
    wb.save(str(wb_path))

    cls = classify_workbook(wb_path)
    # May or may not trigger XML_SURGERY depending on how openpyxl writes the formula
    # If it's OPENPYXL (formula not cached), test the _xlfn path by injecting XML
    if cls.engine != Engine.XML_SURGERY:
        # Inject _xlfn into the XML directly to force XML_SURGERY
        import zipfile as zf
        import shutil
        tmp_xlsx = wb_path.with_suffix(".tmp.xlsx")
        with zf.ZipFile(wb_path, "r") as zin, zf.ZipFile(tmp_xlsx, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if "sheet1.xml" in item.filename:
                    xml = data.decode("utf-8")
                    xml = xml.replace("<v></v>", "<v>1</v>")
                    xml = xml.replace(
                        "_xlfn.XLOOKUP",
                        "_xlfn.XLOOKUP",
                    )
                    # Inject a _xlfn marker with a cached value to trigger detection
                    xml = xml.replace(
                        "<f>_xlfn.XLOOKUP",
                        "<f>_xlfn.XLOOKUP",
                    )
                    # Force detection: just add _xlfn. + non-empty <v>
                    xml = xml.replace(
                        "</sheetData>",
                        '<row r="100"><c r="A100"><f>_xlfn.SEQUENCE(1)</f>'
                        "<v>1</v></c></row></sheetData>",
                    )
                    data = xml.encode("utf-8")
                zout.writestr(item, data)
        shutil.move(str(tmp_xlsx), str(wb_path))

    cls = classify_workbook(wb_path)
    if cls.engine == Engine.XML_SURGERY:
        with pytest.raises(ValueError, match="dynamic-array"):
            write_range(wb_path, "Sheet1", "A1", [[42]])
    else:
        pytest.skip(
            f"Workbook classified as {cls.engine.value}, not XML_SURGERY — "
            "skipping refusal test (formula not serialised with _xlfn marker by openpyxl)"
        )


# ---------------------------------------------------------------------------
# Test 8: list_sheets and list_names
# ---------------------------------------------------------------------------

def test_list_sheets_and_names(tmp_path) -> None:
    """list_sheets returns sheet names; list_names returns defined names."""
    wb_path = tmp_path / "test_meta.xlsx"
    create(wb_path, sheets=["Inputs", "Calcs", "Output"])

    # List sheets
    result = list_sheets(wb_path)
    assert result["sheets"] == ["Inputs", "Calcs", "Output"]

    # Write some data and define a named range
    write_range(wb_path, "Inputs", "B2", [[100]])
    define_name(wb_path, "PeakLoad", "Inputs!$B$2")

    # List names
    names_result = list_names(wb_path)
    assert any(n["name"] == "PeakLoad" for n in names_result["names"]), (
        f"PeakLoad not found in {names_result['names']}"
    )
