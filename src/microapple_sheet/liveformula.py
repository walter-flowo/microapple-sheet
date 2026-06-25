"""liveformula.py — live-formula enforcement and hardcoded-value audit.

Phase 3 — IMPLEMENTED.

Live-formula doctrine (from Rules_Python_Openpyxl_Scripts.md):
  - Every calculated cell must hold a live formula that references its inputs.
  - A "formula" that is a bare number is a defect (it rots silently when inputs change).
  - Outputs are always formulas; inputs are the only literals.

Functions
---------
validate_formula(formula)
    Gate called by write_formula / excel_write_formula before any write.
    Returns (is_valid, reason) — False when formula is a bare number or fails
    basic sanity checks.

audit_hardcoded(path, sheet, range_)
    Lint a workbook for cells that look calculated but are literals.
    Column-peer heuristic: in whole-sheet mode, a numeric literal in a column
    where at least one peer cell is a formula is flagged as hardcoded.
    Strict mode (range_ given): every non-formula numeric cell in the range is
    flagged regardless of column composition.

compose_linked_formula(template, refs)
    Substitute {name} placeholders with A1 coordinates captured from Cell.coordinate.
    Implements Rules_Python_Openpyxl_Scripts.md Rule #1/#4.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# validate_formula
# ---------------------------------------------------------------------------

# A1-style cell references: 1-3 letters + 1+ digits.
# Excludes pure function calls (words followed by '(').
_RE_A1_REF = re.compile(r"[A-Za-z]{1,3}\d+")

# Named range token: 3+ chars of word-chars (letters/digits/_), NOT followed by '('
# This catches names like MY_RANGE or PeakLoad.
_RE_NAMED_RANGE = re.compile(r"[A-Za-z_][A-Za-z0-9_]{2,}(?!\s*\()")

# Quoted string literals to strip before reference detection
_RE_QUOTED = re.compile(r'"[^"]*"')

# Excel function-call pattern: WORD( — remove these to avoid mistaking function
# names for named ranges.
_RE_FUNC_CALL = re.compile(r"[A-Za-z_][A-Za-z0-9_]*\s*\(")


def validate_formula(formula: str) -> tuple[bool, str | None]:
    """Validate that *formula* is a genuine Excel formula, not a bare value.

    Algorithm:
      1. Strip the leading '='.
      2. Remove quoted string literals.
      3. Check for A1-style cell references ([A-Za-z]{1-3}[0-9]+).
      4. Check for named-range tokens (word chars, NOT followed by '(').
      5. If none found → formula is a constant expression → return (False, warning).
      6. Otherwise → (True, None).

    Args:
        formula: The candidate formula string (with or without leading '=').

    Returns:
        (is_valid: bool, reason: str | None)   reason is None when valid.
    """
    if not formula or not formula.strip():
        return (False, "Formula is empty.")

    body = formula.strip()
    if body.startswith("="):
        body = body[1:]

    # Remove quoted string literals
    body_no_strings = _RE_QUOTED.sub("", body)

    # Check for A1-style refs
    if _RE_A1_REF.search(body_no_strings):
        return (True, None)

    # Remove function calls (WORD followed by paren) to isolate named ranges
    body_no_funcs = _RE_FUNC_CALL.sub("", body_no_strings)

    # Check for named range tokens (3+ char word not followed by '(')
    if _RE_NAMED_RANGE.search(body_no_funcs):
        return (True, None)

    return (
        False,
        "Formula contains no cell or range references — appears to be a constant expression. "
        "Use a cell reference (e.g. '=B2+B3') instead of a literal value (e.g. '=42').",
    )


# ---------------------------------------------------------------------------
# audit_hardcoded
# ---------------------------------------------------------------------------

def _is_formula_cell(value: Any) -> bool:
    """Return True if the cell value is a formula string starting with '='."""
    return isinstance(value, str) and value.startswith("=")


def _is_numeric_literal(value: Any) -> bool:
    """Return True if the cell value is a numeric literal (int or float)."""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def audit_hardcoded(
    path: str | Path,
    sheet: str,
    range_: str | None = None,
) -> dict[str, Any]:
    """Lint a workbook for cells that look calculated but contain literal values.

    Two modes:

    **Whole-sheet mode (range_ is None):**
      - Iterates all cells in the used range, skipping row 1 (assumed header).
      - For each column, categorises data-row cells as formula, numeric_literal,
        or other (empty/text/bool).
      - A column with ≥1 formula cell is a "formula column".
      - Any numeric literal in a formula column is flagged as hardcoded.
      - Pure-input columns (no formula cells) are NOT flagged.

    **Strict mode (range_ given, e.g. "H3:J5"):**
      - Every non-formula numeric cell in the named range is flagged, regardless
        of the surrounding column composition.

    The workbook is opened with data_only=False so formula strings are visible.

    Args:
        path:   Absolute path to the workbook.
        sheet:  Sheet name to audit.
        range_: Optional A1-notation range for strict-mode audit.

    Returns:
        {
          "sheet": str,
          "scope": str,
          "hardcoded": list[{"cell": str, "value": float|int, "reason": str}],
          "formula_cells": int,
          "input_literal_cells": int,
          "verdict": "clean" | "hardcoded_found",
        }
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), data_only=False)
    try:
        ws = wb[sheet]

        if range_ is not None:
            # ── Strict mode: every numeric literal in the specified range is flagged ──
            hardcoded: list[dict[str, Any]] = []
            formula_count = 0
            input_count = 0

            cells_in_range = ws[range_]
            # ws[range_] returns a tuple of tuples (rows × cols)
            # Handle both single-row/single-cell and multi-row cases
            if not isinstance(cells_in_range, tuple):
                cells_in_range = ((cells_in_range,),)
            elif cells_in_range and not isinstance(cells_in_range[0], tuple):
                # Single row → wrap
                cells_in_range = (cells_in_range,)

            for row in cells_in_range:
                for cell in row:
                    v = cell.value
                    if _is_formula_cell(v):
                        formula_count += 1
                    elif _is_numeric_literal(v):
                        hardcoded.append({
                            "cell": cell.coordinate,
                            "value": v,
                            "reason": "numeric literal in audited range (strict mode)",
                        })

            return {
                "sheet": ws.title,
                "scope": range_,
                "hardcoded": hardcoded,
                "formula_cells": formula_count,
                "input_literal_cells": input_count,
                "verdict": "hardcoded_found" if hardcoded else "clean",
            }

        else:
            # ── Whole-sheet mode: column-peer heuristic (totals-row aware) ──
            # Step 1: detect totals/summary rows.
            # Step 2: classify data rows vs totals rows; determine formula columns
            #         using DATA rows only (so a SUM at the bottom of an input
            #         column does NOT cause the inputs above it to be flagged).
            # Step 3: handle totals rows separately — formula cells are OK;
            #         numeric literals in totals rows are flagged as hardcoded totals.

            min_row = ws.min_row or 1
            max_row = ws.max_row or 1
            min_col = ws.min_column or 1
            max_col = ws.max_column or 1

            # ── Step 1: identify totals rows ──
            _TOTALS_TEXT = frozenset(
                {"total", "subtotal", "sum", "average", "grand total"}
            )
            _TOTALS_FORMULA_PREFIXES = (
                "sum(", "subtotal(", "average(", "count(", "counta(",
                "min(", "max(", "sumif(", "sumifs(",
            )

            totals_row_numbers: set[int] = set()
            for row in ws.iter_rows(
                min_row=max(min_row, 2),
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    v = cell.value
                    # Text cell whose content matches a totals keyword
                    if isinstance(v, str) and not v.startswith("="):
                        if v.strip().lower() in _TOTALS_TEXT:
                            totals_row_numbers.add(cell.row)
                            break
                    # Formula starting with one of the totals-aggregate functions
                    elif _is_formula_cell(v):
                        body = v[1:].lower().lstrip()  # strip leading '='
                        if any(body.startswith(p) for p in _TOTALS_FORMULA_PREFIXES):
                            totals_row_numbers.add(cell.row)
                            break

            # ── Step 2: collect data-row and totals-row cells ──
            # col_formulas_data[c]  = formula cells in DATA rows of column c
            # col_literals_data[c]  = literal cells in DATA rows of column c
            # totals_cells           = all cells in totals rows (formula or literal)
            col_formulas_data: dict[int, list[Any]] = {}
            col_literals_data: dict[int, list[Any]] = {}
            totals_formula_cells: list[Any] = []
            totals_literal_cells: list[Any] = []

            for row in ws.iter_rows(
                min_row=max(min_row, 2),  # skip row 1 (header)
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    v = cell.value
                    if cell.row in totals_row_numbers:
                        if _is_formula_cell(v):
                            totals_formula_cells.append(cell)
                        elif _is_numeric_literal(v):
                            totals_literal_cells.append(cell)
                    else:
                        # Data row
                        c = cell.column
                        if _is_formula_cell(v):
                            col_formulas_data.setdefault(c, []).append(cell)
                        elif _is_numeric_literal(v):
                            col_literals_data.setdefault(c, []).append(cell)

            # ── Step 3: build hardcoded list ──
            hardcoded_list: list[dict[str, Any]] = []

            # Data rows: flag literals only in formula columns (determined from data rows)
            input_literal_count = 0
            for c, literal_cells in col_literals_data.items():
                if c in col_formulas_data:
                    # Formula column (based on data rows) — flag each literal
                    for cell in literal_cells:
                        hardcoded_list.append({
                            "cell": cell.coordinate,
                            "value": cell.value,
                            "reason": "numeric literal in formula column",
                        })
                else:
                    # Pure-input column in data rows — do NOT flag
                    input_literal_count += len(literal_cells)

            # Totals rows: numeric literals are always flagged; formulas are fine
            for cell in totals_literal_cells:
                hardcoded_list.append({
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "reason": "hardcoded total — should be an aggregate formula",
                })

            formula_cell_count = (
                sum(len(v) for v in col_formulas_data.values())
                + len(totals_formula_cells)
            )

            return {
                "sheet": ws.title,
                "scope": "used_range",
                "hardcoded": hardcoded_list,
                "formula_cells": formula_cell_count,
                "input_literal_cells": input_literal_count,
                "verdict": "hardcoded_found" if hardcoded_list else "clean",
            }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# compose_linked_formula
# ---------------------------------------------------------------------------

def compose_linked_formula(template: str, refs: dict[str, str]) -> str:
    """Build a live formula by substituting {name} placeholders with cell coordinates.

    Example:
        template = '={peak}*{fraction}'
        refs     = {'peak': 'B4', 'fraction': 'C2'}
        → '=B4*C2'

    Args:
        template: Formula template with {name} placeholders.
        refs:     Map of placeholder → A1 coordinate (Cell.coordinate values).

    Returns:
        The resolved formula string.
    """
    return template.format(**refs)
