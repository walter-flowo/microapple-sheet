"""test_rule0_regression.py — Rule-0 safety regression.

Proves that apply_edits / recalc leave ≥ as many cached formula values as before
the edit, AND that the edited input value propagates through dependents.

Measured values on macOS / LO 26.2:
  populated_v BEFORE any edit:       330   (10 literal + 320 formula-cached)
  naive openpyxl save (the bug):      10   ← 97% loss (formula caches wiped)
  safe path (openpyxl + LO recalc):  330   ← matches baseline, edit applied

Run standalone:
    cd <project_root>
    .venv/bin/python3 tests/test_rule0_regression.py

Run via pytest:
    pytest tests/test_rule0_regression.py -v
"""
from __future__ import annotations

import re
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl

_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
_SRC = _ROOT / "src"
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

import pytest

from microapple_sheet.engine import Engine, classify_workbook
from microapple_sheet.libreoffice import apply_edits, recalc, _find_soffice

# Skip the LO-dependent tests when soffice is not installed on this host.
_lo_installed = _find_soffice() is not None
_lo_skip = pytest.mark.skipif(
    not _lo_installed,
    reason="LibreOffice (soffice) not installed on this host",
)


# ---------------------------------------------------------------------------
# Workbook builder + helpers
# ---------------------------------------------------------------------------

def _build_computed_workbook(path: Path) -> None:
    """Build a workbook with 2 sheets and 10×32 formula cells.

    Inputs sheet : A1:A10 = 1..10 (literal values)
    Calcs  sheet : 10 rows × 32 columns of cross-sheet formulas
        Col A  : =Inputs!A<n>*<n>
        Col B  : =A<n>+Inputs!A<n>
        Cols C-AF: =SUM(A<n>) variants

    After LO recalc: 320 formula-cached <v> values + 10 literal = 330 total.
    """
    wb = openpyxl.Workbook()
    inp = wb.active
    inp.title = "Inputs"
    for r in range(1, 11):
        inp.cell(row=r, column=1, value=r)

    calc = wb.create_sheet("Calcs")
    for r in range(1, 11):
        for c in range(1, 33):
            if c == 1:
                formula = f"=Inputs!A{r}*{r}"
            elif c == 2:
                formula = f"=A{r}+Inputs!A{r}"
            else:
                formula = f"=SUM(A{r})"
            calc.cell(row=r, column=c, value=formula)

    wb.save(str(path))


def _count_populated_v(xlsx_path: Path) -> int:
    """Count all <v>…</v> elements with non-empty content across all worksheets."""
    pattern = re.compile(r"<v>[^<\s][^<]*</v>")
    count = 0
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        for name in zf.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                xml = zf.read(name).decode("utf-8", errors="replace")
                count += len(pattern.findall(xml))
    return count


def _read_cell_v(xlsx_path: Path, sheet_index: int, cell_ref: str) -> str | None:
    """Return the cached <v> of a cell from raw xlsx XML (sheet_index 1-based)."""
    sheet_file = f"xl/worksheets/sheet{sheet_index}.xml"
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        if sheet_file not in zf.namelist():
            return None
        xml = zf.read(sheet_file).decode("utf-8", errors="replace")
    pat = re.compile(
        r'<c\s[^>]*r="' + re.escape(cell_ref) + r'"[^>]*>.*?<v>([^<]*)</v>',
        re.DOTALL,
    )
    m = pat.search(xml)
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Core helpers (not directly collected by pytest — prefixed with _)
# ---------------------------------------------------------------------------

def _baseline_after_lo_recalc() -> int:
    """Build workbook → LO recalc → count populated <v>. Returns count (must be >= 300)."""
    with tempfile.TemporaryDirectory(prefix="rule0_test_") as tmpdir:
        wb_path = Path(tmpdir) / "rule0_test.xlsx"
        _build_computed_workbook(wb_path)

        # Before recalc: formula cells have empty <v>; only Inputs literal cells
        # contribute. _count_populated_v counts ALL <v>, so expect < 50 here.
        before = _count_populated_v(wb_path)
        assert before < 50, (
            f"Expected few populated <v> before recalc (literal cells only), got {before}."
        )

        result = recalc(wb_path)
        assert result["recalc_ok"], f"recalc failed: {result['detail']}"

        after = _count_populated_v(wb_path)
        assert after >= 300, (
            f"LO recalc produced only {after} populated <v> (expected >= 300). "
            "Recalculation may have failed silently."
        )
        print(f"  [PASS] baseline_after_lo_recalc: {before} → {after} populated <v>")
        return after


def _check_naive_openpyxl_drops_cached_values(baseline: int) -> None:
    """Build + recalc → naive openpyxl save → count drops to ~10. Proves the bug."""
    with tempfile.TemporaryDirectory(prefix="rule0_naive_") as tmpdir:
        wb_path = Path(tmpdir) / "rule0_naive.xlsx"
        _build_computed_workbook(wb_path)
        recalc(wb_path)

        before = _count_populated_v(wb_path)
        assert before >= 300

        # THE BUG: openpyxl load→save drops ALL formula-cached <v> values
        wb = openpyxl.load_workbook(str(wb_path))
        wb.save(str(wb_path))
        wb.close()

        after = _count_populated_v(wb_path)
        assert after < 50, (
            f"Expected naive openpyxl to collapse cached values (<50), but got {after}. "
            "openpyxl behaviour may have changed."
        )
        print(f"  [PASS] naive_openpyxl_drops_cached_values: {before} → {after} (bug confirmed)")


def _check_apply_edits_preserves_cached_values(baseline: int) -> None:
    """apply_edits (openpyxl + LO recalc) must restore >= baseline AND update dependents."""
    with tempfile.TemporaryDirectory(prefix="rule0_safe_") as tmpdir:
        wb_path = Path(tmpdir) / "rule0_safe.xlsx"
        _build_computed_workbook(wb_path)
        recalc(wb_path)

        before = _count_populated_v(wb_path)
        assert before >= 300

        result = apply_edits(
            wb_path,
            edits=[{"sheet": "Inputs", "cell": "A1", "type": "value", "data": 99}],
        )
        assert result["recalc_ok"], f"apply_edits failed: {result['detail']}"
        assert result["ops_applied"] == 1
        assert result["engine"] == Engine.LIBREOFFICE.value, (
            f"Expected engine=libreoffice, got {result['engine']}"
        )

        after = _count_populated_v(wb_path)
        assert after >= baseline, (
            f"Cached values dropped: {before} → {after} (expected >= {baseline}). "
            "Rule-0 violation."
        )

        input_v = _read_cell_v(wb_path, sheet_index=1, cell_ref="A1")
        assert input_v == "99", f"Inputs!A1 expected '99', got {input_v!r}"

        calcs_a1 = _read_cell_v(wb_path, sheet_index=2, cell_ref="A1")
        assert calcs_a1 is not None, "Calcs!A1 has no cached <v> after apply_edits"
        assert float(calcs_a1) == 99.0, (
            f"Calcs!A1 expected 99.0, got {calcs_a1!r}. Dependent cell not updated."
        )

        print(
            f"  [PASS] apply_edits_preserves_cached_values: "
            f"{before} → {after} (>= {baseline}), "
            f"Inputs!A1={input_v}, Calcs!A1={calcs_a1}"
        )


def _check_classify_workbook_routing() -> None:
    """classify_workbook returns correct Engine for fresh, computed, and nonexistent."""
    with tempfile.TemporaryDirectory(prefix="rule0_classify_") as tmpdir:
        tmp = Path(tmpdir)

        # Nonexistent → OPENPYXL
        cls_new = classify_workbook(tmp / "new.xlsx")
        assert cls_new.engine == Engine.OPENPYXL, (
            f"Expected OPENPYXL for nonexistent, got {cls_new.engine}"
        )
        assert not cls_new.has_cached_values
        assert not cls_new.has_dynamic_arrays

        # Fresh openpyxl build (no cached formula <v>) → OPENPYXL
        fresh_path = tmp / "fresh.xlsx"
        _build_computed_workbook(fresh_path)
        cls_fresh = classify_workbook(fresh_path)
        assert cls_fresh.engine == Engine.OPENPYXL, (
            f"Expected OPENPYXL for fresh workbook, got {cls_fresh.engine}"
        )
        assert not cls_fresh.has_cached_values
        assert cls_fresh.sheets == ["Inputs", "Calcs"]

        # LO-recalced workbook (has formula-cached <v>) → LIBREOFFICE
        recalc(fresh_path)
        cls_comp = classify_workbook(fresh_path)
        assert cls_comp.engine == Engine.LIBREOFFICE, (
            f"Expected LIBREOFFICE for computed workbook, got {cls_comp.engine}"
        )
        assert cls_comp.has_cached_values
        assert cls_comp.disk_hash is not None

        print(
            "  [PASS] classify_workbook_routing: "
            "nonexistent→OPENPYXL, fresh→OPENPYXL, computed→LIBREOFFICE"
        )


# ---------------------------------------------------------------------------
# Standalone runner
# ---------------------------------------------------------------------------

def main() -> None:
    print("Rule-0 Regression Suite")
    print("─" * 50)

    failed: list[str] = []

    def run(fn, *args):
        try:
            return fn(*args)
        except Exception as e:
            print(f"  [FAIL] {fn.__name__}: {e}")
            failed.append(fn.__name__)
            return None

    baseline = run(_baseline_after_lo_recalc)
    if baseline is None:
        baseline = 300

    run(_check_naive_openpyxl_drops_cached_values, baseline)
    run(_check_apply_edits_preserves_cached_values, baseline)
    run(_check_classify_workbook_routing)

    print("─" * 50)
    if failed:
        print(f"FAIL — {len(failed)} test(s) failed: {', '.join(failed)}")
        sys.exit(1)
    else:
        print("PASS — all Rule-0 regression tests passed")


if __name__ == "__main__":
    main()


# ---------------------------------------------------------------------------
# pytest-discoverable wrappers (no param-taking, no return values)
# ---------------------------------------------------------------------------

@_lo_skip
def test_rule0_baseline() -> None:
    """Baseline: LO recalc populates >= 300 cached formula <v> values."""
    baseline = _baseline_after_lo_recalc()
    assert baseline >= 300


@_lo_skip
def test_rule0_naive_drop() -> None:
    """Naive openpyxl save collapses cached values to < 50 (bug proof)."""
    _check_naive_openpyxl_drops_cached_values(300)


@_lo_skip
def test_rule0_safe_path() -> None:
    """Safe path (openpyxl + LO recalc) restores all cached values and updates dependents."""
    _check_apply_edits_preserves_cached_values(300)


@_lo_skip
def test_rule0_classify() -> None:
    """classify_workbook routes correctly for nonexistent / fresh / computed."""
    _check_classify_workbook_routing()
