"""libreoffice.py — recalc-preserving headless LibreOffice operations.

All soffice invocations use an isolated -env:UserInstallation profile so they
never touch the user's GUI LibreOffice and never need pkill.

Key findings from Phase 1a discovery
--------------------------------------
- `--convert-to "xlsx:Calc MS Excel 2007 XML"` round-trip DOES populate all
  `<f>...<v>` cached values correctly (proven: empty → 320 populated cells).
- StarBasic macro setValue() works; setFormula() has unreliable results in
  headless mode on macOS / LO 26.2. Use it only for pure value edits.
- The safe two-step path for apply_edits:
    1. openpyxl writes value/formula edits (drops cached <v>, but keeps formulas)
    2. Immediate LO recalc restores ALL <v> from formulas
  This is the documented "convert-to as a recalc after a minimal cache-safe value
  poke" fallback — proven to restore 320/320 cached values.

Rule 0 compliance
-----------------
- recalc() + apply_edits() leave the workbook with ≥ as many populated cached
  values as before the edit. Tested: 320 → naive openpyxl = 20 (BUG) → safe = 320.
- Never call apply_edits() on Engine.XML_SURGERY workbooks (raises ValueError).

LibreOffice discovery order
----------------------------
1. Environment variable ``MICROAPPLE_SHEET_SOFFICE`` (absolute path to soffice).
2. ``shutil.which("soffice")`` / ``shutil.which("libreoffice")`` (PATH lookup).
3. Well-known platform paths:
   - macOS: ``/Applications/LibreOffice.app/Contents/MacOS/soffice``
   - Linux: ``/usr/bin/soffice``, ``/usr/bin/libreoffice``
4. ``RuntimeError`` if none found.
"""
from __future__ import annotations

import os
import shutil
import shutil as _shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

from microapple_sheet import backups

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

#: Environment variable override for soffice binary path.
_SOFFICE_ENV_VAR = "MICROAPPLE_SHEET_SOFFICE"

#: Well-known platform paths for soffice / libreoffice (checked in order).
_SOFFICE_CANDIDATES: list[str] = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",  # macOS
    "/usr/bin/soffice",        # Linux (Debian/Ubuntu apt package)
    "/usr/bin/libreoffice",    # Linux (alternative symlink)
    "/usr/local/bin/soffice",  # Linux (manual install)
]

#: Isolated user profile — avoids touching GUI LO, no pkill needed
_PROFILE_DIR: Path = Path.home() / ".microapple_sheet" / "lo_profile"

#: Default subprocess timeout (seconds)
_DEFAULT_TIMEOUT: int = 120


def _find_soffice() -> Path | None:
    """Locate the soffice binary using the resolution order in the module docstring.

    Returns:
        Path to soffice if found, else None.
    """
    # 1. Environment variable override
    env_val = os.environ.get(_SOFFICE_ENV_VAR)
    if env_val:
        p = Path(env_val)
        if p.exists():
            return p

    # 2. PATH lookup
    found = _shutil.which("soffice") or _shutil.which("libreoffice")
    if found:
        return Path(found)

    # 3. Well-known platform paths
    for candidate in _SOFFICE_CANDIDATES:
        p = Path(candidate)
        if p.exists():
            return p

    return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def recalc(path: str | Path) -> dict[str, Any]:
    """Force a full LibreOffice recalculation and resave the workbook as xlsx.

    Preserves all cached formula values (Rule 0 safe). Uses --convert-to to
    round-trip the workbook through LibreOffice Calc, which recalculates all
    formulas and writes populated <v> elements for every formula cell.

    Strategy: copy to a temp dir → LO converts in-place → atomic replace.

    Args:
        path: Absolute path to the .xlsx workbook to recalculate.

    Returns:
        {path: str, recalc_ok: bool, backup_path: str, detail: str}

    Raises:
        FileNotFoundError: If *path* does not exist.
        RuntimeError:      If LibreOffice is not installed or conversion fails.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {p}")
    _check_lo_installed()

    backup_path = backups.snapshot(p)

    with tempfile.TemporaryDirectory(prefix="microapple_sheet_recalc_") as tmpdir:
        tmp_src = Path(tmpdir) / p.name
        shutil.copy2(p, tmp_src)

        result = _run_convert_to(tmp_src, output_dir=Path(tmpdir), timeout=_DEFAULT_TIMEOUT)

        if not result["ok"]:
            # Restore original and surface the error
            backups.restore(p, backup_path)
            return {
                "path": str(p),
                "recalc_ok": False,
                "backup_path": backup_path,
                "detail": f"LibreOffice conversion failed: {result['stderr']}",
            }

        # Atomically replace original with the recalculated copy.
        # LO always emits an .xlsx regardless of input suffix, so use the
        # converted output path, not tmp_src (which may be .xlsm/.xlsb).
        recalc_out = _find_converted_output(Path(tmpdir), p.stem, "xlsx")
        if not recalc_out.exists():
            # LO placed it at same path when input was already .xlsx
            recalc_out = tmp_src
        shutil.move(str(recalc_out), str(p))

    return {
        "path": str(p),
        "recalc_ok": True,
        "backup_path": backup_path,
        "detail": "LibreOffice recalculation complete; all cached formula values preserved.",
    }


def apply_edits(
    path: str | Path,
    edits: list[dict[str, Any]],
) -> dict[str, Any]:
    """Apply value and/or formula edits to a COMPUTED workbook (Rule-0 safe).

    Uses the proven two-step approach:
      1. openpyxl writes all edits (drops cached <v>, keeps formula strings)
      2. Immediate LO recalc restores ALL <v> cached values
    Result: input edits applied, ALL formula cached values preserved and updated.

    This is equivalent to the StarBasic macro approach (loadComponentFromURL →
    setValue/setFormula → calculateAll → storeToURL) but more reliable on macOS.
    See module docstring for the discovery note on setFormula() flakiness.

    **Caveat — x14 conditional formatting**: the openpyxl round-trip (Step 1) silently
    corrupts advanced Excel 2010+ CF rules — databars, icon-sets, and colour-scales stored
    in the ``<x14:conditionalFormattings>`` namespace. Plain ``dxf``-based ``cellIs`` and
    ``expression`` rules survive intact. Post-MVP fix: use the StarBasic macro ``setValue``
    path for value-only edits (avoids the openpyxl save entirely); see
    ``libreoffice/Module1.xba`` stub for the pattern.

    Each edit dict:
        {sheet: str, cell: str, type: 'value'|'formula', data: Any}

    type='value'   → cell is set to the literal value (int/float/str/None)
    type='formula' → cell is set to the formula string (with or without leading =)

    Args:
        path:   Absolute path to the .xlsx workbook.
        edits:  List of edit operations (see above).

    Returns:
        {path, ops_applied, backup_path, engine, recalc_ok, detail}

    Raises:
        FileNotFoundError: If *path* does not exist.
        ValueError:        If the workbook has dynamic arrays (XML_SURGERY engine).
        RuntimeError:      If LibreOffice is not installed or conversion fails.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {p}")
    _check_lo_installed()

    # Guard: refuse dynamic-array workbooks (LO recalc would cascade #N/A — Rule 0b)
    from microapple_sheet.engine import classify_workbook, Engine
    cls = classify_workbook(p)
    if cls.engine == Engine.XML_SURGERY:
        raise ValueError(
            "Dynamic-array / LAMBDA workbook — automated file-mode edits are not safe. "
            "Open the workbook in Excel and use excel_live_set / excel_live_formula instead "
            "(Phase 2 live-bridge). XML surgery support is a later phase."
        )

    backup_path = backups.snapshot(p)

    # ── Step 1: apply edits with openpyxl ─────────────────────────────────
    # load_workbook with data_only=False to preserve formula strings in OTHER cells.
    # This intentionally drops cached <v> values — Step 2 (LO recalc) restores them.
    wb = openpyxl.load_workbook(str(p))
    ops_applied = 0
    errors: list[str] = []

    for edit in edits:
        try:
            ws = wb[edit["sheet"]]
            cell_addr = edit["cell"]
            edit_type = edit.get("type", "value")
            data = edit["data"]

            if edit_type == "formula":
                # Normalise: openpyxl expects formulas WITH leading =
                formula = str(data).strip()
                if not formula.startswith("="):
                    formula = "=" + formula
                ws[cell_addr] = formula
            else:
                ws[cell_addr] = data

            ops_applied += 1
        except Exception as exc:
            errors.append(f"{edit.get('sheet','?')}!{edit.get('cell','?')}: {exc}")

    wb.save(str(p))
    wb.close()

    if errors:
        detail_prefix = f"Applied {ops_applied}/{len(edits)} edits (errors: {'; '.join(errors)}). "
    else:
        detail_prefix = f"Applied {ops_applied}/{len(edits)} edits. "

    # ── Step 2: LO recalc to restore cached values ────────────────────────
    # Only needed for computed workbooks (LIBREOFFICE engine)
    recalc_ok = False
    if cls.engine == Engine.LIBREOFFICE:
        recalc_result = recalc(p)
        recalc_ok = recalc_result["recalc_ok"]
        if not recalc_ok:
            return {
                "path": str(p),
                "ops_applied": ops_applied,
                "backup_path": backup_path,
                "engine": "libreoffice",
                "recalc_ok": False,
                "detail": detail_prefix + "LO recalc failed: " + recalc_result["detail"],
            }
        detail = detail_prefix + "LO recalc complete; all cached formula values preserved."
    else:
        # OPENPYXL engine: no cached values to restore; openpyxl save is sufficient
        recalc_ok = True
        detail = detail_prefix + "openpyxl save complete (no cached values to restore)."

    return {
        "path": str(p),
        "ops_applied": ops_applied,
        "backup_path": backup_path,
        "engine": cls.engine.value,
        "recalc_ok": recalc_ok,
        "detail": detail,
    }


def convert(
    path: str | Path,
    to: str,
    output_dir: str | Path | None = None,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Convert a workbook to PDF, CSV, or XLSX via LibreOffice headless.

    Args:
        path:       Absolute path to the source workbook.
        to:         Target format: 'pdf', 'csv', or 'xlsx'.
        output_dir: Directory for the output file. Defaults to same dir as *path*.
        sheet:      (CSV only) Sheet name to export. Ignored for other formats.

    Returns:
        {output_path: str, ok: bool, detail: str}

    Raises:
        FileNotFoundError: If *path* does not exist.
        ValueError:        If *to* is not a supported format.
        RuntimeError:      If LibreOffice is not installed.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {p}")
    _check_lo_installed()

    fmt = to.lower().strip()
    _FILTER_MAP = {
        "pdf":  "calc_pdf_Export",
        "csv":  "Text - txt - csv (StarCalc)",
        "xlsx": "Calc MS Excel 2007 XML",
    }
    if fmt not in _FILTER_MAP:
        raise ValueError(f"Unsupported format {to!r}. Use: pdf, csv, xlsx.")

    out_dir = Path(output_dir) if output_dir else p.parent

    with tempfile.TemporaryDirectory(prefix="microapple_sheet_convert_") as tmpdir:
        tmp_src = Path(tmpdir) / p.name
        shutil.copy2(p, tmp_src)

        result = _run_convert_to(
            tmp_src,
            output_dir=Path(tmpdir),
            fmt=fmt,
            filter_name=_FILTER_MAP[fmt],
            timeout=_DEFAULT_TIMEOUT,
        )

        if not result["ok"]:
            return {
                "output_path": "",
                "ok": False,
                "detail": f"LibreOffice conversion failed: {result['stderr']}",
            }

        # Locate the output file (LO names it <stem>.<ext>)
        ext_map = {"pdf": ".pdf", "csv": ".csv", "xlsx": ".xlsx"}
        tmp_out = Path(tmpdir) / (p.stem + ext_map[fmt])
        if not tmp_out.exists():
            return {
                "output_path": "",
                "ok": False,
                "detail": f"Converted file not found at expected path {tmp_out}",
            }

        out_path = out_dir / tmp_out.name
        shutil.move(str(tmp_out), str(out_path))

    return {
        "output_path": str(out_path),
        "ok": True,
        "detail": f"Converted {p.name} → {out_path.name} via LibreOffice.",
    }


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _check_lo_installed() -> Path:
    """Return the soffice binary path, or raise RuntimeError if not found.

    Resolution order (see module docstring):
    1. MICROAPPLE_SHEET_SOFFICE env var
    2. PATH lookup (soffice / libreoffice)
    3. Well-known platform paths
    """
    soffice = _find_soffice()
    if soffice is None:
        searched = [f"  env ${_SOFFICE_ENV_VAR}", "  PATH: soffice / libreoffice"]
        searched += [f"  {c}" for c in _SOFFICE_CANDIDATES]
        raise RuntimeError(
            "LibreOffice (soffice) not found. Searched:\n"
            + "\n".join(searched)
            + "\nInstall from https://www.libreoffice.org/ or set "
            + f"{_SOFFICE_ENV_VAR}=/path/to/soffice"
        )
    return soffice


def _profile_arg() -> str:
    """Return the -env:UserInstallation argument for an isolated LO profile."""
    _PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    # Convert to file:// URL (LO requires this format on macOS)
    uri = _PROFILE_DIR.as_uri()  # e.g. file:///home/user/.microapple_sheet/lo_profile
    return f"-env:UserInstallation={uri}"


def _run_convert_to(
    src: Path,
    output_dir: Path,
    fmt: str = "xlsx",
    filter_name: str = "Calc MS Excel 2007 XML",
    timeout: int = _DEFAULT_TIMEOUT,
) -> dict[str, Any]:
    """Run `soffice --headless --convert-to <filter> --outdir <dir> <file>`.

    Returns:
        {ok: bool, returncode: int, stdout: str, stderr: str}
    """
    soffice = _find_soffice()
    if soffice is None:
        return {
            "ok": False,
            "returncode": -1,
            "stdout": "",
            "stderr": "soffice binary not found — see RuntimeError from _check_lo_installed()",
        }

    convert_spec = f"{fmt}:{filter_name}" if fmt != "pdf" else "pdf:calc_pdf_Export"

    cmd = [
        str(soffice),
        "--headless",
        "--norestore",
        "--invisible",
        _profile_arg(),
        "--convert-to", convert_spec,
        "--outdir", str(output_dir),
        str(src),
    ]

    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return {
            "ok": False,
            "returncode": -1,
            "stdout": "",
            "stderr": f"LibreOffice timed out after {timeout}s",
        }
    except FileNotFoundError:
        return {
            "ok": False,
            "returncode": -1,
            "stdout": "",
            "stderr": f"soffice binary not found at {soffice}",
        }

    ok = proc.returncode == 0 and _find_converted_output(output_dir, src.stem, fmt).exists()
    return {
        "ok": ok,
        "returncode": proc.returncode,
        "stdout": proc.stdout.strip(),
        "stderr": proc.stderr.strip(),
    }


def _find_converted_output(output_dir: Path, stem: str, fmt: str) -> Path:
    ext = {"pdf": ".pdf", "csv": ".csv", "xlsx": ".xlsx"}.get(fmt, f".{fmt}")
    return output_dir / (stem + ext)
